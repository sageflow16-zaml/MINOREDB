"""Trade import/export service supporting CSV, Excel, and JSON."""
import csv
import io
import json
from datetime import datetime, timezone
from uuid import UUID, uuid4
from typing import Any
from fastapi import UploadFile, HTTPException, status
from sqlalchemy import select
from sqlalchemy.orm import Session
from src.models.trade import Trade
from src.models.trade_import import TradeImport
from src.models.strategy import Strategy
from src.core.logging import get_logger

logger = get_logger(__name__)

FIELD_MAP = {
    "symbol": "pair",
    "pair": "pair",
    "direction": "direction",
    "entry_price": "entry_price",
    "entry": "entry_price",
    "exit_price": "exit_price",
    "exit": "exit_price",
    "close_price": "exit_price",
    "stop_loss": "stop_loss",
    "sl": "stop_loss",
    "take_profit": "take_profit",
    "tp": "take_profit",
    "risk_percent": "risk_percent",
    "risk": "risk_percent",
    "risk_%": "risk_percent",
    "lot_size": "position_size",
    "position_size": "position_size",
    "size": "position_size",
    "rr": "rr",
    "r_r": "rr",
    "pnl": "pnl",
    "p_l": "pnl",
    "profit": "pnl",
    "commission": "commission",
    "swap": "swap",
    "broker": "broker_name",
    "broker_name": "broker_name",
    "strategy": "strategy",
    "session": "session",
    "timeframe": "timeframe",
    "open_time": "open_time",
    "open": "open_time",
    "close_time": "close_time",
    "closed": "close_time",
    "close": "close_time",
    "tags": "tags",
    "notes": "notes",
    "psychology": "emotion",
    "emotion": "emotion",
    "screenshot": "before_image",
    "screenshot_path": "before_image",
    "before_image": "before_image",
    "result": "result",
    "status": "status",
    "weekly_bias": "weekly_bias",
    "daily_bias": "daily_bias",
    "h4_bias": "h4_bias",
    "liquidity_sweep": "liquidity_sweep",
    "bos": "bos",
    "mss": "mss",
    "order_block": "order_block",
    "fvg": "fvg",
    "asian_session": "asian_session",
    "london_session": "london_session",
    "newyork_session": "newyork_session",
    "dxy": "dxy",
    "us10y": "us10y",
    "us02y": "us02y",
    "news_event": "news_event",
}

NUMERIC_FIELDS = {
    "entry_price", "exit_price", "stop_loss", "take_profit",
    "position_size", "risk_percent", "rr", "pnl", "commission", "swap",
}

BOOLEAN_SESSION_FIELDS = {"asian_session", "london_session", "newyork_session"}
SESSION_NAMES = {"asian", "london", "newyork", "ny", "us", "uk", "europe"}


def _normalise_row(raw: dict[str, Any]) -> dict[str, Any]:
    mapped: dict[str, Any] = {}
    for k, v in raw.items():
        if k is None or v is None:
            continue
        sk = k.strip().lower().replace("-", "_").replace(" ", "_")
        field = FIELD_MAP.get(sk)
        if field is None:
            continue
        if field in NUMERIC_FIELDS:
            try:
                mapped[field] = float(str(v).replace(",", "").replace("$", "").replace("€", ""))
            except (ValueError, TypeError):
                raise ValueError(f"Invalid numeric value '{v}' for {k}")
        elif field in {"open_time", "close_time"}:
            parsed = _parse_datetime(str(v))
            if parsed is not None:
                mapped[field] = parsed.isoformat()
        elif field == "tags":
            if isinstance(v, str):
                mapped[field] = [t.strip() for t in v.split(",") if t.strip()]
            elif isinstance(v, list):
                mapped[field] = [str(t) for t in v]
        elif field == "direction":
            mapped[field] = str(v).upper().strip()
        elif field == "session":
            sv = str(v).strip().lower()
            if sv in SESSION_NAMES:
                if sv in ("ny", "us"):
                    mapped["newyork_session"] = "true"
                elif sv in ("uk", "europe"):
                    mapped["london_session"] = "true"
                else:
                    mapped[f"{sv}_session"] = "true"
        elif field == "strategy":
            mapped["_strategy_name"] = str(v).strip()
        elif field == "broker_name":
            mapped["broker_name"] = str(v).strip()
        else:
            mapped[field] = str(v).strip() if isinstance(v, str) else v
    return mapped


def _parse_datetime(s: str) -> datetime | None:
    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s.strip(), fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def _validate_row(row: dict[str, Any], row_num: int) -> list[str]:
    errors = []
    if not row.get("pair"):
        errors.append(f"Row {row_num}: Symbol is required")
    direction = row.get("direction", "").upper()
    if direction and direction not in ("BUY", "SELL", "B", "S", "LONG", "SHORT"):
        errors.append(f"Row {row_num}: Invalid direction '{row.get('direction')}'")
    if row.get("entry_price") is not None and not isinstance(row["entry_price"], (int, float)):
        errors.append(f"Row {row_num}: Entry price must be numeric")
    return errors


def _normalise_direction(d: str | None) -> str | None:
    if not d:
        return None
    d = d.upper().strip()
    if d in ("B", "LONG", "BUY"):
        return "BUY"
    if d in ("S", "SHORT", "SELL"):
        return "SELL"
    return d


def _ensure_strategy(db: Session, project_id: UUID, name: str) -> UUID | None:
    if not name:
        return None
    stmt = select(Strategy).where(
        Strategy.project_id == project_id,
        Strategy.name == name,
    )
    existing = db.scalar(stmt)
    if existing:
        return existing.id
    s = Strategy(id=uuid4(), project_id=project_id, name=name)
    db.add(s)
    db.flush()
    return s.id


def _find_duplicate(db: Session, project_id: UUID, row: dict[str, Any]) -> Trade | None:
    if not row.get("pair"):
        return None
    filters = [Trade.project_id == project_id, Trade.pair == row["pair"]]
    if row.get("direction"):
        filters.append(Trade.direction == row["direction"])
    if row.get("entry_price") is not None:
        filters.append(Trade.entry_price == row["entry_price"])
    if row.get("open_time"):
        filters.append(Trade.open_time == row["open_time"])
    stmt = select(Trade).where(*filters)
    return db.scalar(stmt)


def _trade_to_dict(t: Trade) -> dict[str, Any]:
    return {
        "id": str(t.id),
        "project_id": str(t.project_id),
        "pair": t.pair,
        "direction": t.direction,
        "entry_price": t.entry_price,
        "exit_price": t.exit_price,
        "stop_loss": t.stop_loss,
        "take_profit": t.take_profit,
        "position_size": t.position_size,
        "risk_percent": t.risk_percent,
        "rr": t.rr,
        "pnl": t.pnl,
        "commission": t.commission,
        "swap": t.swap,
        "result": t.result,
        "status": t.status,
        "broker_name": t.broker_name,
        "timeframe": t.timeframe,
        "open_time": t.open_time.isoformat() if t.open_time else None,
        "close_time": t.close_time.isoformat() if t.close_time else None,
        "tags": t.tags,
        "weekly_bias": t.weekly_bias,
        "daily_bias": t.daily_bias,
        "h4_bias": t.h4_bias,
        "emotion": t.emotion,
        "notes": t.notes,
        "before_image": t.before_image,
        "after_image": t.after_image,
        "asian_session": t.asian_session,
        "london_session": t.london_session,
        "newyork_session": t.newyork_session,
        "dxy": t.dxy,
        "us10y": t.us10y,
        "us02y": t.us02y,
        "news_event": t.news_event,
        "liquidity_sweep": t.liquidity_sweep,
        "bos": t.bos,
        "mss": t.mss,
        "order_block": t.order_block,
        "fvg": t.fvg,
        "created_at": t.created_at.isoformat() if t.created_at else None,
    }


# ── Parser ──

ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".json"}


def parse_file(file: UploadFile) -> list[dict[str, Any]]:
    ext = _get_ext(file.filename)
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"Unsupported format '{ext}'. Use .csv, .xlsx, or .json")
    raw = file.file.read()
    if not raw or len(raw) == 0:
        raise HTTPException(status_code=400, detail="Empty file")
    if ext == ".csv":
        return _parse_csv(raw)
    elif ext == ".xlsx":
        return _parse_xlsx(raw)
    else:
        return _parse_json(raw)


def _get_ext(filename: str | None) -> str:
    if not filename:
        return ".csv"
    return "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ".csv"


def _parse_csv(raw: bytes) -> list[dict[str, Any]]:
    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def _parse_xlsx(raw: bytes) -> list[dict[str, Any]]:
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row_vals in rows[1:]:
        row_dict = {}
        for i, h in enumerate(headers):
            val = row_vals[i] if i < len(row_vals) else None
            if val is not None:
                row_dict[h] = val
        result.append(row_dict)
    return result


def _parse_json(raw: bytes) -> list[dict[str, Any]]:
    try:
        data = json.loads(raw.decode("utf-8-sig"))
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON file")
    if isinstance(data, dict):
        data = data.get("trades", data.get("data", [data]))
    if not isinstance(data, list):
        raise HTTPException(status_code=400, detail="JSON must contain an array of trades")
    return data


# ── Preview ──


def preview_import(db: Session, project_id: UUID, file: UploadFile) -> dict[str, Any]:
    raw_rows = parse_file(file)
    rows = []
    valid_count = 0
    duplicate_count = 0
    error_count = 0
    for i, raw in enumerate(raw_rows):
        row_num = i + 1
        try:
            mapped = _normalise_row(raw)
        except ValueError as e:
            rows.append({"row_number": row_num, "data": {}, "errors": [str(e)], "is_duplicate": False, "duplicate_of": None})
            error_count += 1
            continue
        errors = _validate_row(mapped, row_num)
        duplicate = _find_duplicate(db, project_id, mapped) if not errors else None
        row = {
            "row_number": row_num,
            "data": mapped,
            "errors": errors,
            "is_duplicate": duplicate is not None,
            "duplicate_of": str(duplicate.id) if duplicate else None,
        }
        rows.append(row)
        if errors:
            error_count += 1
        elif duplicate:
            duplicate_count += 1
        else:
            valid_count += 1

    imp = TradeImport(
        project_id=project_id,
        filename=file.filename or "import",
        format=_get_ext(file.filename).lstrip("."),
        total_rows=len(raw_rows),
        preview_data={"rows": rows},
    )
    db.add(imp)
    db.commit()
    db.refresh(imp)

    return {
        "import_id": str(imp.id),
        "filename": imp.filename,
        "format": imp.format,
        "total_rows": imp.total_rows,
        "valid_rows": valid_count,
        "duplicate_rows": duplicate_count,
        "error_rows": error_count,
        "rows": rows,
        "created_at": imp.created_at.isoformat(),
    }


# ── Confirm Import ──


def confirm_import(
    db: Session, project_id: UUID, import_id: UUID, duplicate_strategy: str = "skip",
) -> dict[str, Any]:
    imp = db.get(TradeImport, import_id)
    if not imp or imp.project_id != project_id:
        raise HTTPException(status_code=404, detail="Import session not found")
    if imp.status != "pending":
        raise HTTPException(status_code=400, detail="Import already processed")

    preview = imp.preview_data or {"rows": []}
    rows = preview.get("rows", [])
    imported = 0
    updated = 0
    skipped = 0
    failed = 0
    details: list[dict[str, Any]] = []

    try:
        for row in rows:
            rn = row.get("row_number", 0)
            if row.get("errors"):
                failed += 1
                details.append({"row": rn, "action": "failed", "errors": row["errors"]})
                continue
            if row.get("is_duplicate") and row.get("duplicate_of"):
                if duplicate_strategy == "skip":
                    skipped += 1
                    details.append({"row": rn, "action": "skipped", "reason": "duplicate"})
                    continue
                elif duplicate_strategy == "update":
                    trade = db.get(Trade, UUID(row["duplicate_of"]))
                    if trade:
                        _apply_row(trade, row["data"])
                        trade.updated_at = datetime.now(timezone.utc)
                        db.add(trade)
                        updated += 1
                        details.append({"row": rn, "action": "updated", "trade_id": str(trade.id)})
                        continue
            trade = Trade(id=uuid4(), project_id=project_id)
            _apply_row(trade, row["data"])
            db.add(trade)
            imported += 1
            details.append({"row": rn, "action": "imported", "trade_id": str(trade.id)})

        imp.status = "completed"
        imp.imported_count = imported
        imp.updated_count = updated
        imp.skipped_count = skipped
        imp.failed_count = failed
        imp.completed_at = datetime.now(timezone.utc)
        imp.error_rows = {"details": details}
        db.commit()
    except Exception as exc:
        db.rollback()
        imp.status = "failed"
        db.commit()
        logger.error("Import %s failed: %s", import_id, exc, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Import failed: {exc}")

    return {
        "import_id": str(import_id),
        "status": "completed",
        "total_rows": imp.total_rows,
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
        "failed": failed,
        "details": details,
    }


def _apply_row(trade: Trade, data: dict[str, Any]) -> None:
    for field in (
        "pair", "direction", "entry_price", "exit_price", "stop_loss",
        "take_profit", "position_size", "risk_percent", "rr", "pnl",
        "commission", "swap", "result", "status", "broker_name", "timeframe",
        "weekly_bias", "daily_bias", "h4_bias",
        "liquidity_sweep", "bos", "mss", "order_block", "fvg",
        "asian_session", "london_session", "newyork_session",
        "dxy", "us10y", "us02y", "news_event",
        "emotion", "notes", "before_image", "after_image", "tags",
    ):
        if field in data:
            setattr(trade, field, data[field])
    if data.get("open_time"):
        ot = data["open_time"]
        if isinstance(ot, str):
            trade.open_time = _parse_datetime(ot)
        else:
            trade.open_time = ot
    if data.get("close_time"):
        ct = data["close_time"]
        if isinstance(ct, str):
            trade.close_time = _parse_datetime(ct)
        else:
            trade.close_time = ct
    direction = data.get("direction")
    if direction:
        trade.direction = _normalise_direction(direction)
    strategy_name = data.get("_strategy_name")
    if strategy_name and not trade.strategy_id:
        trade.strategy_id = _ensure_strategy(
            db_bare(trade), trade.project_id, strategy_name,
        )


def db_bare(obj) -> Session:
    return Session.object_session(obj)


# ── Export ──


def export_trades(
    db: Session,
    project_id: UUID,
    fmt: str,
    filters: dict[str, Any] | None = None,
) -> tuple[str, str, str]:
    stmt = select(Trade).where(Trade.project_id == project_id).order_by(Trade.created_at.desc())
    if filters:
        if filters.get("ids"):
            stmt = stmt.where(Trade.id.in_([UUID(i) for i in filters["ids"]]))
        if filters.get("date_from"):
            stmt = stmt.where(Trade.created_at >= filters["date_from"])
        if filters.get("date_to"):
            stmt = stmt.where(Trade.created_at <= filters["date_to"])
        if filters.get("symbol"):
            stmt = stmt.where(Trade.pair == filters["symbol"])
        if filters.get("strategy_id"):
            stmt = stmt.where(Trade.strategy_id == UUID(filters["strategy_id"]))
        if filters.get("result"):
            stmt = stmt.where(Trade.result == filters["result"])
        if filters.get("status"):
            stmt = stmt.where(Trade.status == filters["status"])
        if filters.get("broker"):
            stmt = stmt.where(Trade.broker_name == filters["broker"])
        if filters.get("tags"):
            stmt = stmt.where(Trade.tags.has_any(filters["tags"]))

    trades = list(db.scalars(stmt).all())
    data = [_trade_to_dict(t) for t in trades]
    fmt = fmt.lower()
    if fmt == "json":
        content = json.dumps({"trades": data}, indent=2, default=str)
        media = "application/json"
        ext = "json"
    elif fmt == "xlsx":
        content, media, ext = _export_xlsx(data)
    else:
        content, media, ext = _export_csv(data)
    filename = f"trades_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"
    return content, media, filename


def _export_csv(data: list[dict[str, Any]]) -> tuple[str, str, str]:
    if not data:
        return "", "text/csv", "csv"
    headers = list(data[0].keys())
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row in data:
        clean = {k: v if not isinstance(v, list) else ",".join(v) for k, v in row.items()}
        writer.writerow(clean)
    return output.getvalue(), "text/csv", "csv"


def _export_xlsx(data: list[dict[str, Any]]) -> tuple[bytes, str, str]:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trades"
    if not data:
        return b"", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"
    headers = list(data[0].keys())
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(data, 2):
        for ci, h in enumerate(headers, 1):
            val = row.get(h)
            if isinstance(val, list):
                val = ",".join(val)
            cell = ws.cell(row=ri, column=ci, value=val)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"


def get_import_history(db: Session, project_id: UUID) -> list[TradeImport]:
    stmt = (
        select(TradeImport)
        .where(TradeImport.project_id == project_id)
        .order_by(TradeImport.created_at.desc())
        .limit(50)
    )
    return list(db.scalars(stmt).all())
